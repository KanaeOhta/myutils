from functools import reduce

import openpyxl
from xlsxwriter.workbook import Workbook


class ExcelMixin:

    def _flatten(self, items, key, val, pref):
        if isinstance(val, dict):
            return {**items, **self.flatten(val, pref + key + '.')}
        elif isinstance(val, list):
            if val and isinstance(val[0], dict):
                for i, dic in enumerate(val):
                    return {**items, **self.flatten(dic, pref + key + '.' + str(i) + '.')}
            else:
                return {**items, pref + key + '.0': ','.join(val)}
        else:
            return {**items, pref + key: val}


    def flatten(self, d, pref=''):
        return reduce(lambda new_d, kv: self._flatten(new_d, *kv, pref), d.items(), {})


    def json_to_excel(self, js_dicts, output_file):
        keys = None
        with Workbook(output_file) as wb:
            sh = wb.add_worksheet()
            for row, js in enumerate(js_dicts, 1):
                if not keys:
                    keys = tuple(js.keys())
                    sh.set_column(0, len(keys)-1, 20)
                    [sh.write_string(0, col, str(key)) for col, key in enumerate(keys)]
                for col, key in enumerate(keys):
                    if type(js[key]) == int:
                        sh.write(row, col, js[key])
                    else:
                        sh.write_string(row, col, str(js[key]))


    def read(self, excel_file, settings):
        wb = openpyxl.load_workbook(excel_file, read_only=True)
        sh = wb.worksheets[0]
        self.keys = sh[1]
        for row in sh.iter_rows(**settings):
            yield row
        wb.close()


    def _unflatten(self, new_dic, keys, val):
        if len(keys) == 1:
            new_dic[keys[0]] = val
            return
        if keys[1].isnumeric():
            i = int(keys[1])
            if keys[0] not in new_dic:
                new_dic[keys[0]] = [{}]
            if len(new_dic[keys[0]]) < i + 1:
                new_dic[keys[0]].append({})
            self._unflatten(new_dic[keys[0]][i], keys[2:], val)
        else:
            if keys[0] not in new_dic:
                new_dic[keys[0]] = {}
            self._unflatten(new_dic[keys[0]], keys[1:], val)

    
    def unflatten(self, dic):
        new_dic = {}
        for key, val in dic.items():
            keys = key.split('.')
            if keys[-1].isnumeric():
                keys = keys[:-1]
                val = val.split(',') if val else list()
            self._unflatten(new_dic, keys, val)
        return new_dic


    def excel_to_json(self, excel_file):
        settings = {'min_row': 2}
        for row in self.read(excel_file, settings):
            dic = {k.value: v.value if v.value is not None \
                else '' for k, v in zip(self.keys, row)}
            yield self.unflatten(dic)
    

    def write(self, records, output_file):
        """records must be namedtuple or dict."""
        keys = None
        with Workbook(output_file) as wb:
            sh = wb.add_worksheet()
            self.set_format(wb)
            for row, record in enumerate(records, 1):
                if keys is None:
                    keys = self.set_header(record)
                    [sh.write_string(0, col, str(key)) for col, key in enumerate(keys)]
                    sh.set_column(0, len(keys)-1, 30)
                for col, key in enumerate(keys):
                    self._write(sh, row, col, key, record)
        return keys
    

    def _write(self, sh, row, col, key, record):
        """
        Override this method in subclasses to use write method.
        """
        pass

    def set_header(self, record):
        """
        Override this method in subclasses to use write method.
        example: return record._fields or return recods.keys() 
        """
        pass


    def set_format(self, wb):
        """
        Override this method in subclasses to set excel sheet format.
        example: 
            self.fmt = wb.add_format()
            self.fmt.set_font_color('blue')
        """
        pass

    