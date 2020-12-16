from collections import namedtuple
from datetime import datetime
from functools import partial
import errno
import itertools
import json
import os
import re

import openpyxl
from xlsxwriter.workbook import Workbook


class NoMoreRecord(Exception):
    pass


class ExtensionError(Exception):
    pass


class JsonFile:

    def __init__(self, file):
        self.file = file


    def __iter__(self):
        with open(self.file, 'r', encoding='utf-8') as f:
            for dic in json.load(f):
                yield dic


    def output(self, records, indent):
        with open(self.file, 'w', encoding='utf-8') as f:
            json.dump(
                list(records), 
                f, 
                ensure_ascii=False,
                indent=indent
            )


Cell = namedtuple('Cell', 'key idx value')


class ExcelSheet:

    MAIN = 'main'

    def __init__(self, sheet):
        self.sheet = sheet


    def set_keys(self, *args):
        raise NotImplementedError()


class WritingSheet(ExcelSheet):

    URL = r"(https?|ftp)(:\/\/[-_\.!~*\'()a-zA-Z0-9;\/?:\@&=\+$,%#]+)"


    def __init__(self, sheet, keys, row=0, index=''):
        super().__init__(sheet)
        self._row = row
        self._index = index
        self.pattern = re.compile(WritingSheet.URL)
        self.set_keys(keys)
    

    def set_keys(self, keys):
        self.keys = {}
        for col, key in enumerate(keys, 1):
            self.keys[key] = col
            self.write(0, col, key)


    def column(self, key):
        return self.keys.get(key)


    def row(self, index):
        if self._index != index:
            self._index = index
            self._row += 1
        return self._row


    def write(self, row, col, value, index=''):
        if index:
            self.sheet.write(row, 0, index)
        if not col:
            pass
        if type(value) in {float, int}:
            self.sheet.write_number(row, col, value)
        elif isinstance(value, list) or value is None:
            self.sheet.write_blank(row, col, None)
        elif type(value) == bool:
            self.sheet.write_boolean(row, col, value)
        elif self.pattern.match(value):
            self.sheet.write_url(row, col, value)
        elif isinstance(value, str):
            self.sheet.write_string(row, col, value)
        else:
            self.sheet.write(row, col, value)
       

class ReadingSheet(ExcelSheet):

    def __init__(self, sheet):
        super().__init__(sheet)
        self.title = self.sheet.title
        self.keys = self.set_keys()
        self.max_col = len(self.keys) + 1
        self.max_row = self.sheet.max_row
        self.data_types = tuple(data_type for data_type in self.get_type())
        self.row = 2
  

    def is_empty(self, cell_value):
        return cell_value is None or not str(cell_value).strip()    
    

    def set_keys(self):
        row = self.sheet[1]
        return tuple(cell.value for cell in row if not self.is_empty(cell.value))


    def get_type(self):
        """For each column, yield data type.
        """
        for col in self.sheet.iter_cols(
                min_row=2, max_row=self.max_row, min_col=2, max_col=self.max_col):
            data_type = None  
            for cell in col:
                if not self.is_empty(cell.value):
                    data_type = type(cell.value)
                    # 0.0 is entered into cell as 0.   
                    if data_type != int:
                        break
            yield data_type


    def _read(self, cells):
        for cell, key, data_type in zip(cells, self.keys, self.data_types):
            if not self.is_empty(val := cell.value):
                val = data_type(val)
                if data_type == str:
                    # Convert '_x000D_' to '\r'
                    val = openpyxl.utils.escape.unescape(val)
            yield key, val


    def read(self, serial, delimiter):
        """If Rows have the same left number of Indexes(column A) split with hyphen,
           they are on record. For example, 1, 1-1, 1-2, 1-1-1, 1-1-2.   
        """
        if self.row > self.max_row:
            raise NoMoreRecord() 
        min_row = self.row
        for row in self.sheet.iter_rows(
                min_row=min_row, max_row = self.max_row, min_col=1, max_col=self.max_col):
            a_cell = row[0]
            # when formatted cell is found out of data area
            if self.is_empty(a_cell.value):
                raise NoMoreRecord()
            idx = a_cell.value
            if idx.split(delimiter)[0] != serial:
                return
            self.row += 1
            record = {(key, idx): val for key, val in self._read(row[1:])}
            # If all cells in a row are empty and sheet is Main, value is empty list.
            if all(self.is_empty(val) for val in record.values()) and self.title != ExcelSheet.MAIN:
                yield {(self.title, f'{idx}'): list()}
            else:
                yield record 


def file_check(path, ext):
    if not os.path.isfile(path):
        raise FileNotFoundError(
            errno.ENOENT, os.strerror(errno.ENOENT), path)
    if os.path.splitext(path)[-1] != f'.{ext}':
        raise ExtensionError('{} file is not selected.'.format(ext))


class Convert:
    """A mixin class to provide functions to flatten or 
       unflatten dict.
       DOT and HYPEN are immutable. Can be accessed with self.
    """
    DOT = '.'
    HYPHEN = '-'

    def get_file_path(self, original_path, ext):
        str_date = datetime.now().strftime('%Y%m%d%H%M%S')
        path = '{}_{}{}'.format(os.path.splitext(original_path)[0],
            str_date, ext)
        return path


    def serialize(self, dic, idx, pref=''):
        """Flatten dict. 
        """
        for key, val in dic.items():
            if isinstance(val, dict):
                yield from self.serialize(val, 
                    idx, f'{pref}{key}{self.DOT}')
            elif isinstance(val, list):
                if val:
                    for i, list_val in enumerate(val):
                        if isinstance(list_val, dict):
                            yield from self.serialize(list_val, 
                                f'{idx}{self.HYPHEN}{i}', f'{pref}{key}{self.DOT}')
                        else:
                            yield from self.serialize({f'{pref}{key}{self.HYPHEN}{i}' : list_val}, idx)
                else:
                    yield f'{pref}{key}{self.HYPHEN}{str(0)}', idx, val
            else:
                yield f'{pref}{key}', idx, val


    def parse_json(self, dic, group, pref=''):
        """Return a pair of sheet name and column name.
        """
        for key, val in dic.items():
            if isinstance(val, dict):
                yield from self.parse_json(
                    val, group, f'{pref}{key}{self.DOT}')
            elif isinstance(val, list):
                if val:
                    for i, list_val in enumerate(val):
                        if isinstance(list_val, dict):
                            yield from self.parse_json(
                                list_val, f'{pref}{key}', f'{pref}{key}{self.DOT}')
                        else:
                            yield from self.parse_json(
                                {f'{pref}{key}{self.HYPHEN}{i}' : list_val}, group)
                else:
                    yield group, f'{pref}{key}{self.HYPHEN}{str(0)}'
            else:
                yield group, f'{pref}{key}'


    def set_obj_to_list(self, li, idx, obj):
        """obj must be list or dict.
        """
        try:
            li[idx]
        except IndexError:
            li.append(obj())
        return li[idx]


    def set_obj_to_dict(self, dic, key, obj):
        """obj must be list or dict.
        """
        if key not in dic:
            dic[key] = obj()
        return dic[key]

    
    def set_nested_list(self, li, split_keys):
        """split_keys, li: list
            Return nested list or li.
        """
        for split_key in split_keys:
            li = self.set_obj_to_list(li, int(split_key), list)
        return li


    def _deserialize(self, new_dic, keys, idxes, val):
        if not idxes: # the first level
            if len(keys) == 1:
                if self.HYPHEN in keys[0]: # Value is list
                    split_keys = keys[0].split(self.HYPHEN)
                    # If split_keys[1:-1] is not empty, li is nested list
                    li = self.set_nested_list(
                        self.set_obj_to_dict(new_dic, split_keys[0], list), split_keys[1:-1])
                    if val: 
                        li.append(val)
                else: # Value is not list
                    new_dic[keys[0]] = val
            else: # Nested dict
                dic = self.set_obj_to_dict(new_dic, keys[0], dict)
                self._deserialize(dic, keys[1:], idxes, val)
        else: # Dict in list
            split_keys = keys[0].split(self.HYPHEN)
            # If split_keys[1:-1] is not empty, li is nested list
            li = self.set_nested_list(
                self.set_obj_to_dict(new_dic, split_keys[0], list), split_keys[1:])
            dic = self.set_obj_to_list(li, idxes[0], dict)
            self._deserialize(dic, keys[1:], idxes[1:], val)


    def deserialize(self, dic):
        new_dic = {}
        for (key_str, idx_str), val in dic.items():
            keys =key_str.split(self.DOT)
            idxes = [int(idx) for idx in idx_str.split(self.HYPHEN)]
            self._deserialize(new_dic, keys, idxes[1:], val)
        return new_dic


    def replace(self, _replace, obj):
        """Particular characters in key are replaced. 
           _replace: function 
        """
        if isinstance(obj, dict):
            return {_replace(key): self.replace(_replace, val) for key, val in obj.items()}
        elif isinstance(obj, list):
            return [self.replace(_replace, item) for item in obj]
        else:
            return obj


    def _flatten_list(self, li):
        for item in li:
            if isinstance(item, list):
                yield from self._flatten_list(item)
            else:
                if isinstance(item, dict):
                    yield item

    
    def _find(self, dic, split_keys):
        for k, v in dic.items():
            if len(split_keys) == 1 and k == split_keys[0]:
                if isinstance(v, dict):
                    yield v
                elif isinstance(v, list):
                    yield from self._flatten_list(v)
            elif isinstance(v, dict):
                yield from self._find(v, split_keys[1:])
            elif isinstance(v, list):
                for list_val in self._flatten_list(v):
                    yield from self._find(list_val, split_keys[1:])

    
    def replace_selected_keys(self, replacement, dic):
        """replacement: {old_key: new_key, ...}
           If dic is {'a': {'b': {'c_c': 5}}, 'd': 6} and replacement is {'a.b.c_c': 'c-c'},
           dic will be changed to {'a': {'b': {'c-c': 5}}, 'd': 6}
        """
        for old_keys, new_key in replacement.items():
            split_keys = old_keys.split('.')
            # split_keys[-1] will be replaced
            old_key = split_keys[-1]
            if len(split_keys) == 1 and old_key in dic:
                dic[new_key] = dic.pop(old_key)
                continue 
            for found in self._find(dic, split_keys[:-1]):
                found[new_key] = found.pop(old_key)
        return dic


    def separate(self, keys):
        """Separate key connected with DOT as below.
           If keys id 'AA.BB.CC', 
           ('AA', 'AA), ('AA.BB', 'BB) and ('AA.BB.CC', 'CC) are returned. 
        """
        li = [key.split(self.HYPHEN)[0] for key in keys.split(self.DOT)]
        for i, key in enumerate(li, 1):
            yield '.'.join(li[:i]), li[:i][-1] 


class ToExcel(Convert):
    """Write data in json file to worksheets in Excel.
    """

    def __init__(self, json_file):
        json_file = os.path.abspath(json_file)
        file_check(json_file, 'json')
        self.json = JsonFile(json_file)
        self.set_replace_table(
            {self.HYPHEN: '_', self.DOT: '_'})
        self.sheet_format = {}
        self.sheets = None
       

    def set_replace_table(self, replacement): 
        table = str.maketrans(replacement)
        self._replace = lambda x: x.translate(table)
    
    
    def set_sheet_format(self):
        """Create dict {column name: Column namedtuple}
        """
        if not self.sheet_format:
            for dic in self.json:
                for group, key in self.parse_json(
                        self.replace(self._replace, dic), group=ExcelSheet.MAIN): 
                    if key not in self.sheet_format.keys():
                        self.sheet_format[key] = group


    def get_selected_records(self, keys):
        for i, dic in enumerate(self.json, 1):
            yield (Cell(key, idx, val) for key, idx, val \
                in self.serialize(self.replace(self._replace, dic), str(i)) if key in keys)


    def partial_convert(self, *keys):
        self.set_sheet_format()
        self.sheet_format = {key: val for key, val \
            in self.sheet_format.items() if key in keys}
        records = iter(self.get_selected_records(keys))
        self.output(records)
        self.sheet_format = {}


    def get_records(self):
        for i, dic in enumerate(self.json, 1):
            yield (Cell(key, idx, val) for key, idx, val \
                in self.serialize(self.replace(self._replace, dic), str(i)))


    def convert(self):
        self.set_sheet_format()
        records = iter(self.get_records())
        self.output(records)
        self.sheet_format = {}


    def set_sheets(self, wb):
        """Create dict {sh_name: WritingSheet instance}
           If a key is equal to sh_name directory connected to HYPHEN,
           the key is empty object list.
        """
        self.sheets = {}
        # Change sh_name if key is sh_name connected to directory HYPHEN 
        for sh_name in set(self.sheet_format.values()):
            for key in self.sheet_format.keys():
                if key == f'{sh_name}-0':
                    self.sheet_format[key] = sh_name
        for sh_name in self.sheet_format.values():
            if sh_name not in self.sheets:
                self.sheets[sh_name] = WritingSheet(
                    wb.add_worksheet(sh_name),
                    sorted(key for key, val in self.sheet_format.items() \
                        if val == sh_name and key != f'{sh_name}-0')
                )

  
    def write(self, sh_name, cell):
        """Get column and row numbers to write value to worksheet.
        """
        sheet = self.sheets[sh_name]
        col = sheet.column(cell.key)
        row = sheet.row(cell.idx)
        sheet.write(row, col, cell.value, cell.idx)


    def output(self, records):
        excel_file = self.get_file_path(self.json.file, '.xlsx')
        with Workbook(excel_file) as wb:
            self.set_sheets(wb)
            for record in records:
                for cell in record:
                    sh_name = self.sheet_format.get(cell.key)
                    self.write(sh_name, cell)


class FromExcel(Convert):
    """Read data on worksheets in Excel to output to json file.
       It is necessary to use Excel file created using ToExcel instance.
    """

    def __init__(self, excel_file):
        excel_file = os.path.abspath(excel_file)
        file_check(excel_file, 'xlsx')
        self.output_file = self.get_file_path(excel_file, '.json')
        self.wb = openpyxl.load_workbook(excel_file)
        self.sheets = None
        
        
    def set_sheets(self):
        if self.sheets is None:
            self.sheets = tuple(ReadingSheet(sh) for sh \
                in self.wb if sh.cell(row=2, column=1).value)
    

    def convert(self, indent=None, replacement=None):
        self.set_sheets()
        if replacement:
            _replace = partial(self.replace_selected_keys, replacement)
            records = (_replace(record) for record in self.read())
        else:
            records = (record for record in self.read())
        self.output(records, indent)
        self.wb.close()

    
    def read(self):
        for i in itertools.count(1):
            dic = {}
            try:
                for sheet in self.sheets:
                    for record in sheet.read(str(i), self.HYPHEN):
                        dic = {**dic, **record}
                yield self.deserialize(dic)         
            except NoMoreRecord:
                break


    def output(self, records, indent):
        json_file = JsonFile(self.output_file) 
        json_file.output(records, indent)